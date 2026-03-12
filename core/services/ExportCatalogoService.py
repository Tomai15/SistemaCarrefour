from __future__ import annotations

import json
import os
import logging
import random
import re
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from datetime import datetime
from html import unescape

import requests
from django.conf import settings
from django.db import close_old_connections
from openpyxl import Workbook

from core.models import TareaCatalogacion, SellerVtex

logger: logging.Logger = logging.getLogger(__name__)

CANTIDAD_WORKERS = 100
TAMANIO_PAGINA = 200
REINTENTOS_MAXIMOS = 3
ESPERA_ENTRE_REINTENTOS = 2  # segundos base para backoff exponencial

# Las 52 columnas del export oficial + extras utiles
COLUMNAS_EXPORT = [
    'EAN', 'ACTIVO', 'FOTO', 'CATALOGADO',
    '_IDSKU', '_NombreSku', '_ActivarSKUSiEsPosible', '_SkuActivo', '_EANSKU',
    '_Altura', '_AlturaReal', '_Anchura', '_AnchuraReal',
    '_Longitud', '_LongitudReal', '_Peso', '_PesoReal',
    '_UnidadMedida', '_MultiplicadorUnidad', '_CodigoReferenciaSKU',
    '_ValorFidelidad', '_FechaEstimadaLlegada', '_CodigoFabricante',
    '_IDProducto', '_NombreProducto', '_DescripcionCortaProducto',
    '_ProductoActivo', '_CodigoReferenciaProducto', '_MostrarEnSitio',
    '_LinkTexto', '_DescripcionProducto', '_FechaLanzamientoProducto',
    '_PalabrasClave', '_TituloSitio', '_DescripcionMetaTag',
    '_IDProveedor', '_MostrarSinStock', '_Kit',
    '_IDDepartamento', '_NombreDepartamento', '_IDCategoria', '_NombreCategoria',
    '_IDMarca', '_Marca', '_PesoVolumetrico', '_CondicionComercial',
    '_Tiendas', '_Accesorios', '_Similares', '_Sugerencias',
    '_ShowTogether', '_Adjunto',
    # Extras utiles fuera del export oficial
    'Motivo', 'Precio', 'Stock',
]

COLUMNAS_IMAGENES = [
    '_IDProducto', '_NombreProducto', '_IDSKU', '_NombreSku',
    'ID de imagen', 'Nombre de imagen', 'URL de imagen',
    'Texto de imagen', 'Etiqueta', 'CodigoReferenciaSKU',
]


@dataclass
class _ContextoVtex:
    """Agrupa datos de conexion a VTEX (seller, marketplace, sesiones HTTP)."""
    seller: SellerVtex
    headers_seller: dict
    url_base_seller: str
    headers_marketplace: dict
    url_base_marketplace: str
    sales_channels_filtro: list[int]
    session_marketplace: requests.Session = field(default_factory=requests.Session)
    session_seller: requests.Session = field(default_factory=requests.Session)


class ExportCatalogoService:

    INTERVALO_PROGRESO = 200

    def __init__(self) -> None:
        self._lock = threading.Lock()
        self._progreso_pendiente: int = 0
        # Estado en memoria — se escribe a archivo temporal, no a DB
        self._estado: str = ''
        self._logs: str = ''
        self._progreso_actual: int = 0
        self._progreso_total: int = 0
        self._ruta_temporal: str = ''

    # -- Metodo principal -----------------------------------------------------

    def ejecutar(
        self, tarea: TareaCatalogacion, seller_id: int,
        sales_channels_filtro: list[int] | None = None,
        incluir_precio_stock: bool = True,
    ) -> None:
        """Exporta el catalogo completo de un seller con las 52 columnas del export VTEX.

        Pipeline por fases:
        1. Obtener detalles de todos los SKUs
        2. Obtener productos, categorias y marcas unicos
        3. (Opcional) Obtener precio y stock de cada SKU
        4. Construir filas y generar Excel
        """
        self._iniciar_estado_temporal(tarea)
        contexto = None
        try:
            contexto = self._inicializar(tarea, seller_id, sales_channels_filtro)
            if contexto is None:
                return

            total_fases = 4 if incluir_precio_stock else 3
            inicio_total = time.time()

            # Fase 0: Obtener todos los SKU IDs
            self._log(tarea, "Obteniendo listado de SKU IDs...")
            t0 = time.time()
            sku_ids = self._obtener_todos_los_sku_ids(tarea, contexto)
            if not sku_ids:
                self._log(tarea, "No se encontraron SKU IDs para este seller.")
                self._actualizar_estado(tarea, TareaCatalogacion.Estado.COMPLETADO)
                return

            total_skus = len(sku_ids)
            self._log(tarea, f"{total_skus} SKU IDs encontrados ({time.time() - t0:.1f}s)")

            # Fase 1: Fetch todos los SKU details
            self._log(tarea, f"[1/{total_fases}] Obteniendo detalles de {total_skus} SKUs...")
            t0 = time.time()
            self._set_progreso(total_skus, 0)

            detalles_skus = self._fase_fetch_skus(tarea, sku_ids, contexto)

            skus_ok = sum(1 for d in detalles_skus if d is not None)
            skus_error = total_skus - skus_ok
            msg_fase1 = f"[1/{total_fases}] Completado: {skus_ok} SKUs obtenidos"
            if skus_error:
                msg_fase1 += f", {skus_error} con error"
            msg_fase1 += f" ({time.time() - t0:.1f}s)"
            self._log(tarea, msg_fase1)

            # Extraer IDs unicos de productos
            ids_productos = set()
            for ds in detalles_skus:
                if ds is None:
                    continue
                pid = ds.get('ProductId')
                if pid:
                    ids_productos.add(str(pid))

            # Fase 2: Fetch productos unicos (categorias y marcas salen del SKU, sin requests extra)
            self._log(tarea, f"[2/{total_fases}] Obteniendo {len(ids_productos)} productos...")
            t0 = time.time()
            self._set_progreso(len(ids_productos), 0)

            productos = self._fase_fetch_productos(tarea, ids_productos, contexto)

            self._log(tarea, f"[2/{total_fases}] Completado ({time.time() - t0:.1f}s)")

            # Fase 3 (opcional): Pricing + Stock
            precios: dict[int, float | None] = {}
            stocks: dict[int, int | None] = {}
            if incluir_precio_stock:
                self._log(tarea, f"[3/{total_fases}] Obteniendo precio y stock de {total_skus} SKUs...")
                t0 = time.time()
                self._set_progreso(total_skus * 2, 0)
                precios, stocks = self._fase_fetch_precio_stock(tarea, sku_ids, contexto)

                sin_precio = sum(1 for v in precios.values() if v is None)
                sin_stock = sum(1 for v in stocks.values() if v is not None and v <= 0)
                self._log(tarea, f"[3/{total_fases}] Completado: {sin_precio} sin precio, {sin_stock} sin stock ({time.time() - t0:.1f}s)")

            # Fase final: Construir resultados y generar Excel
            fase_final = total_fases
            self._log(tarea, f"[{fase_final}/{total_fases}] Generando Excel con {total_skus} filas...")
            t0 = time.time()
            resultados = self._construir_resultados(
                sku_ids, detalles_skus, productos,
                precios, stocks, contexto, incluir_precio_stock
            )

            # Construir filas de imagenes antes de liberar memoria
            filas_imagenes = self._construir_filas_imagenes(sku_ids, detalles_skus)

            # Liberar memoria intermedia (ya no se necesitan)
            del detalles_skus, productos, precios, stocks

            # Contar resumen
            activos = sum(1 for r in resultados if r.get('ACTIVO') == 'SI')
            catalogados = sum(1 for r in resultados if r.get('CATALOGADO') == 'SI')

            # Generar Excel
            self._generar_excel(tarea, resultados, filas_imagenes)
            self._actualizar_estado(tarea, TareaCatalogacion.Estado.COMPLETADO)

            tiempo_total = time.time() - inicio_total
            self._log(tarea, f"Export finalizado: {total_skus} SKUs, {activos} activos, {catalogados} catalogados ({tiempo_total:.0f}s total)")
        finally:
            self._finalizar_estado(tarea)
            if contexto:
                contexto.session_marketplace.close()
                contexto.session_seller.close()

    # -- Inicializacion de contexto -------------------------------------------

    def _inicializar(
        self, tarea: TareaCatalogacion, seller_id: int, sales_channels_filtro: list[int] | None
    ) -> _ContextoVtex | None:
        self._actualizar_estado(tarea, TareaCatalogacion.Estado.PROCESANDO)

        try:
            seller = SellerVtex.objects.get(id=seller_id)
        except SellerVtex.DoesNotExist:
            self._log(tarea, f"Error: No se encontro el seller con ID {seller_id}")
            self._actualizar_estado(tarea, TareaCatalogacion.Estado.ERROR)
            return None

        marketplace = seller.marketplace if seller.marketplace else seller
        self._log(tarea, f"Seller: {seller.nombre} ({seller.account_name})")
        if marketplace != seller:
            self._log(tarea, f"Marketplace madre: {marketplace.nombre} ({marketplace.account_name})")

        headers_base = {'Accept': 'application/json', 'Content-Type': 'application/json'}
        headers_marketplace = {
            **headers_base,
            'X-VTEX-API-AppKey': marketplace.app_key,
            'X-VTEX-API-AppToken': marketplace.app_token,
        }
        headers_seller = {
            **headers_base,
            'X-VTEX-API-AppKey': seller.app_key,
            'X-VTEX-API-AppToken': seller.app_token,
        }

        session_marketplace = requests.Session()
        session_marketplace.headers.update(headers_marketplace)
        session_marketplace.mount('https://', requests.adapters.HTTPAdapter(
            pool_connections=CANTIDAD_WORKERS, pool_maxsize=CANTIDAD_WORKERS + 10
        ))

        session_seller = requests.Session()
        session_seller.headers.update(headers_seller)
        session_seller.mount('https://', requests.adapters.HTTPAdapter(
            pool_connections=CANTIDAD_WORKERS, pool_maxsize=CANTIDAD_WORKERS + 10
        ))

        return _ContextoVtex(
            seller=seller,
            headers_seller=headers_seller,
            url_base_seller=f"https://{seller.account_name}.vtexcommercestable.com.br",
            headers_marketplace=headers_marketplace,
            url_base_marketplace=f"https://{marketplace.account_name}.vtexcommercestable.com.br",
            sales_channels_filtro=sales_channels_filtro if sales_channels_filtro else [1, 3],
            session_marketplace=session_marketplace,
            session_seller=session_seller,
        )

    # -- Obtener todos los SKU IDs (paginado) ---------------------------------

    def _obtener_todos_los_sku_ids(self, tarea: TareaCatalogacion, contexto: _ContextoVtex) -> list[int]:
        todos_los_ids: list[int] = []
        page = 1
        while True:
            url = (
                f"{contexto.url_base_marketplace}"
                f"/api/catalog_system/pvt/sku/stockkeepingunitids"
                f"?page={page}&pagesize={TAMANIO_PAGINA}"
            )
            datos = self._request_con_retry(url, contexto.session_marketplace)
            if datos is None:
                self._log(tarea, f"Error obteniendo pagina {page} de SKU IDs. Abortando.")
                self._actualizar_estado(tarea, TareaCatalogacion.Estado.ERROR)
                return []
            if not datos:
                break
            todos_los_ids.extend(datos)
            logger.info(f"Pagina {page}: {len(datos)} SKU IDs (total: {len(todos_los_ids)})")
            if len(datos) < TAMANIO_PAGINA:
                break
            page += 1
        return todos_los_ids

    # -- Pipeline: Fases de obtencion de datos --------------------------------

    def _fase_fetch_skus(
        self, tarea: TareaCatalogacion, sku_ids: list[int], contexto: _ContextoVtex
    ) -> list[dict | None]:
        """Fase 1: obtiene detalle de cada SKU via stockkeepingunitbyid."""
        detalles_skus: list[dict | None] = [None] * len(sku_ids)
        total = len(sku_ids)
        procesados = 0
        errores = 0
        siguiente_log = max(total // 5, 500)  # Log cada ~20% o cada 500

        def _obtener_detalle_sku(idx_skuid: tuple[int, int]) -> tuple[int, dict | None]:
            idx, sku_id = idx_skuid
            url = f"{contexto.url_base_marketplace}/api/catalog_system/pvt/sku/stockkeepingunitbyid/{sku_id}"
            return idx, self._request_con_retry(url, contexto.session_marketplace, silenciar_404=True)

        with ThreadPoolExecutor(max_workers=CANTIDAD_WORKERS) as pool:
            futuros = {
                pool.submit(_obtener_detalle_sku, (i, sid)): i
                for i, sid in enumerate(sku_ids)
            }
            for futuro in as_completed(futuros):
                try:
                    idx, datos = futuro.result()
                    detalles_skus[idx] = datos
                    if datos is None:
                        errores += 1
                except Exception as e:
                    idx = futuros[futuro]
                    logger.error(f"Error fetch SKU idx {idx}: {e}")
                    errores += 1
                self._incrementar_progreso(tarea)
                procesados += 1
                if procesados % siguiente_log == 0:
                    self._log(tarea, f"  {procesados}/{total} SKUs procesados...")

        self._flush_progreso(tarea)
        return detalles_skus

    def _fase_fetch_productos(
        self, tarea: TareaCatalogacion, product_ids: set[str], contexto: _ContextoVtex
    ) -> dict[str, dict | None]:
        """Fase 2: obtiene datos de productos unicos."""
        productos: dict[str, dict | None] = {}

        def _fetch_producto(pid: str) -> tuple[str, dict | None]:
            url = f"{contexto.url_base_marketplace}/api/catalog/pvt/product/{pid}"
            return pid, self._request_con_retry(url, contexto.session_marketplace, silenciar_404=True)

        with ThreadPoolExecutor(max_workers=CANTIDAD_WORKERS) as pool:
            futuros = [pool.submit(_fetch_producto, pid) for pid in product_ids]
            for f in as_completed(futuros):
                try:
                    pid, datos = f.result()
                    productos[pid] = datos
                except Exception as e:
                    logger.error(f"Error fetch producto: {e}")
                self._incrementar_progreso(tarea)
        self._flush_progreso(tarea)
        return productos

    def _fase_fetch_precio_stock(
        self, tarea: TareaCatalogacion, sku_ids: list[int], contexto: _ContextoVtex
    ) -> tuple[dict[int, float | None], dict[int, int | None]]:
        """Fase 3: obtiene precio y stock en dos pools paralelos."""
        precios: dict[int, float | None] = {}
        stocks: dict[int, int | None] = {}

        def _run_precios():
            with ThreadPoolExecutor(max_workers=CANTIDAD_WORKERS // 2) as pool:
                futuros = {pool.submit(self._obtener_precio, sid, contexto): sid for sid in sku_ids}
                for f in as_completed(futuros):
                    sid = futuros[f]
                    try:
                        precios[sid] = f.result()
                    except Exception:
                        precios[sid] = None
                    self._incrementar_progreso(tarea)

        def _run_stocks():
            with ThreadPoolExecutor(max_workers=CANTIDAD_WORKERS // 2) as pool:
                futuros = {pool.submit(self._obtener_stock, sid, contexto): sid for sid in sku_ids}
                for f in as_completed(futuros):
                    sid = futuros[f]
                    try:
                        stocks[sid] = f.result()
                    except Exception:
                        stocks[sid] = None
                    self._incrementar_progreso(tarea)

        hilo_precios = threading.Thread(target=_run_precios)
        hilo_stocks = threading.Thread(target=_run_stocks)
        hilo_precios.start()
        hilo_stocks.start()
        hilo_precios.join()
        hilo_stocks.join()

        self._flush_progreso(tarea)
        return precios, stocks

    # -- Construir resultados -------------------------------------------------

    def _construir_resultados(
        self, sku_ids: list[int], detalles_skus: list[dict | None],
        productos: dict[str, dict | None],
        precios: dict[int, float | None], stocks: dict[int, int | None],
        contexto: _ContextoVtex, incluir_precio_stock: bool = True,
    ) -> list[dict]:
        """Fase final: combina datos para armar las 52+ columnas por SKU."""
        resultados: list[dict] = []

        for i, sku_id in enumerate(sku_ids):
            datos_sku = detalles_skus[i]
            if datos_sku is None:
                resultados.append(self._resultado_error(sku_id, 'Error al consultar catalogo'))
                continue
            # EAN: buscar primero en ProductSpecifications (FieldId 979), fallback a AlternateIds/Ean
            # PELIGRO, NO CORRESPONDE OBTENER EL EAN DESDE EL PRODUCTO
            #specs = datos_sku.get('ProductSpecifications') or []
            #spec_ean = next((s for s in specs if s.get('FieldId') == 979), None)
            #ean_de_spec = (spec_ean['FieldValues'][0] if spec_ean and spec_ean.get('FieldValues') else '')

            product_id = str(datos_sku.get('ProductId', '') or '')
            ean = datos_sku.get('AlternateIds', {}).get('RefId', '') or datos_sku.get('Ean', '') or ''
            brand_id = str(datos_sku.get('BrandId', '') or '')

            # Producto
            datos_prod = productos.get(product_id) if product_id else None

            # Categoria: primera entrada de ProductCategories (la mas especifica)
            prod_cats = datos_sku.get('ProductCategories') or {}
            cat_items = list(prod_cats.items())
            category_id = cat_items[0][0] if cat_items else ''
            nombre_categoria = cat_items[0][1] if cat_items else ''

            # Departamento: ultima entrada de ProductCategories (la mas general)
            department_id = cat_items[-1][0] if cat_items else ''
            nombre_departamento = cat_items[-1][1] if cat_items else ''

            # Marca (viene directo del SKU)
            nombre_marca = datos_sku.get('BrandName', '') or ''

            # Precio y stock
            precio = precios.get(sku_id)
            stock = stocks.get(sku_id)

            descripcion = (datos_prod or {}).get('Description', '') or ''

            # Calcular columnas
            foto = bool(datos_sku.get('Images', []))
            calidad_ok = self._calcular_calidad(datos_sku, descripcion)
            activo, motivo = self._calcular_activo(
                datos_sku, datos_prod, foto, calidad_ok,
                contexto.sales_channels_filtro, precio, stock,
                incluir_precio_stock,
            )
            # Catalogado: tiene imagen + categoria valida (no vacia, no default, no deshabilitados)
            categorias_invalidas = {'', 'deshabilitados', 'categoria default', 'categoría default'}
            categoria_valida = bool(cat_items) and all(
                str(cat_name).strip().lower() not in categorias_invalidas
                for _, cat_name in cat_items
            )
            catalogado = foto and categoria_valida

            sc_list = datos_sku.get('SalesChannels', [])
            sc_str = ', '.join(str(sc) for sc in sc_list) if sc_list else ''
            dimension = datos_sku.get('Dimension', {}) if isinstance(datos_sku.get('Dimension'), dict) else {}

            resultados.append({
                'EAN': int(ean) if str(ean).isdigit() else ean,
                'ACTIVO': 'SI' if activo else 'NO',
                'FOTO': 'SI' if foto else 'NO',
                'CATALOGADO': 'SI' if catalogado else 'NO',
                '_IDSKU': int(sku_id),
                '_NombreSku': datos_sku.get('NameComplete', '') or datos_sku.get('SkuName', ''),
                '_ActivarSKUSiEsPosible': _si_no(datos_sku.get('ActivateIfPossible')),
                '_SkuActivo': _si_no(datos_sku.get('IsActive')),
                '_EANSKU': int(ean) if str(ean).isdigit() else ean,
                '_Altura': _num(datos_sku.get('Height') or dimension.get('height')),
                '_AlturaReal': _num(datos_sku.get('RealHeight') or dimension.get('realHeight')),
                '_Anchura': _num(datos_sku.get('Width') or dimension.get('width')),
                '_AnchuraReal': _num(datos_sku.get('RealWidth') or dimension.get('realWidth')),
                '_Longitud': _num(datos_sku.get('Length') or dimension.get('length')),
                '_LongitudReal': _num(datos_sku.get('RealLength') or dimension.get('realLength')),
                '_Peso': _num(datos_sku.get('Weight') or dimension.get('weight')),
                '_PesoReal': _num(datos_sku.get('RealWeight') or dimension.get('realWeight')),
                '_UnidadMedida': str(datos_sku.get('MeasurementUnit', '') or ''),
                '_MultiplicadorUnidad': _num(datos_sku.get('UnitMultiplier')),
                '_CodigoReferenciaSKU':  int(ean) if str(ean).isdigit() else ean,
                '_ValorFidelidad': _num(datos_sku.get('RewardValue')),
                '_FechaEstimadaLlegada': str(datos_sku.get('EstimatedDateArrival', '') or ''),
                '_CodigoFabricante': str(datos_sku.get('ManufacturerCode', '') or ''),
                '_IDProducto': str(product_id),
                '_NombreProducto': datos_sku.get('ProductName', ''),
                '_DescripcionCortaProducto': (datos_prod or {}).get('ShortDescription', '') or '',
                '_ProductoActivo': _si_no(datos_sku.get('IsProductActive')),
                '_CodigoReferenciaProducto': str(
                    datos_sku.get('ProductRefId', '') or (datos_prod or {}).get('RefId', '') or ''
                ),
                '_MostrarEnSitio': _si_no((datos_prod or {}).get('IsVisible')),
                '_LinkTexto': str((datos_prod or {}).get('LinkId', '') or ''),
                '_DescripcionProducto': descripcion,
                '_FechaLanzamientoProducto': str((datos_prod or {}).get('ReleaseDate', '') or ''),
                '_PalabrasClave': str((datos_prod or {}).get('KeyWords', '') or ''),
                '_TituloSitio': str((datos_prod or {}).get('Title', '') or ''),
                '_DescripcionMetaTag': str((datos_prod or {}).get('MetaTagDescription', '') or ''),
                '_IDProveedor': str((datos_prod or {}).get('SupplierId', '') or ''),
                '_MostrarSinStock': _si_no((datos_prod or {}).get('ShowWithoutStock')),
                '_Kit': _si_no(datos_sku.get('IsKit')),
                '_IDDepartamento': str(department_id),
                '_NombreDepartamento': nombre_departamento,
                '_IDCategoria': str(category_id),
                '_NombreCategoria': nombre_categoria,
                '_IDMarca': str(brand_id),
                '_Marca': nombre_marca,
                '_PesoVolumetrico': _num(dimension.get('cubicweight') or dimension.get('CubicWeight')),
                '_CondicionComercial': str(datos_sku.get('CommercialConditionId', '') or ''),
                '_Tiendas': sc_str,
                '_Accesorios': '',
                '_Similares': '',
                '_Sugerencias': '',
                '_ShowTogether': '',
                '_Adjunto': '',
                'Motivo': motivo if not activo else '',
                'Precio': precio,
                'Stock': stock,
            })

        return resultados

    # -- Construir filas de imagenes ------------------------------------------

    def _construir_filas_imagenes(
        self, sku_ids: list[int], detalles_skus: list[dict | None],
    ) -> list[dict]:
        """Genera una fila por cada imagen de cada SKU."""
        filas: list[dict] = []
        for i, sku_id in enumerate(sku_ids):
            datos_sku = detalles_skus[i]
            if datos_sku is None:
                continue
            product_id = str(datos_sku.get('ProductId', '') or '')
            product_name = datos_sku.get('ProductName', '') or ''
            sku_name = datos_sku.get('SkuName', '') or ''
            ref_id = datos_sku.get('AlternateIds', {}).get('RefId', '') or ''
            for img in datos_sku.get('Images', []):
                image_url = img.get('ImageUrl', '') or ''
                # Limpiar URL: quitar https:// y query params
                url_sin_params = image_url.split('?')[0]
                url_limpia = url_sin_params.replace('https://', '').replace('http://', '')
                # Nombre de imagen: ultimo segmento de la URL sin extension
                nombre_archivo = ''
                if '/' in url_sin_params:
                    nombre_archivo = url_sin_params.rsplit('/', 1)[-1].rsplit('.', 1)[0]
                filas.append({
                    '_IDProducto': product_id,
                    '_NombreProducto': product_name,
                    '_IDSKU': int(sku_id),
                    '_NombreSku': sku_name,
                    'ID de imagen': img.get('FileId', ''),
                    'Nombre de imagen': nombre_archivo,
                    'URL de imagen': url_limpia,
                    'Texto de imagen': nombre_archivo,
                    'Etiqueta': '',
                    'CodigoReferenciaSKU': ref_id,
                })
        return filas

    # -- Logica de columnas ---------------------------------------------------

    def _calcular_calidad(self, datos_sku: dict, descripcion: str) -> bool:
        product_name = datos_sku.get('ProductName', '')
        tiene_foto = bool(datos_sku.get('Images', []))
        nombre_ok = not product_name.isupper() if product_name else False
        foto_ok = tiene_foto
        descripcion_ok = not descripcion.isupper() if descripcion else True
        categorias = datos_sku.get('ProductCategories', {})
        categorias_invalidas = {'deshabilitados', 'categoria default', 'categoría default'}
        categoria_ok = True
        for cat_name in categorias.values():
            if str(cat_name).strip().lower() in categorias_invalidas:
                categoria_ok = False
                break
        return nombre_ok and foto_ok and descripcion_ok and categoria_ok

    def _calcular_activo(
        self, datos_sku: dict, datos_prod: dict | None, foto: bool, calidad_ok: bool,
        sales_channels_filtro: list[int], precio: float | None, stock: int | None,
        incluir_precio_stock: bool = True,
    ) -> tuple[bool, str]:
        motivos: list[str] = []
        if not foto:
            motivos.append('Sin imagenes')
        if not calidad_ok:
            product_name = datos_sku.get('ProductName', '')
            if product_name and product_name.isupper():
                motivos.append('Nombre todo mayusculas')
            categorias = datos_sku.get('ProductCategories', {})
            categorias_invalidas = {'deshabilitados', 'categoria default', 'categoría default'}
            for cat_name in categorias.values():
                if str(cat_name).strip().lower() in categorias_invalidas:
                    motivos.append(f'Categoria: {cat_name}')
                    break
            if not motivos:
                motivos.append('No catalogado (calidad)')
        if not datos_sku.get('IsActive', False):
            motivos.append('SKU inactivo')
        if not datos_sku.get('IsProductActive', False):
            motivos.append('Producto inactivo')
        if datos_prod and not datos_prod.get('IsVisible', False):
            motivos.append('No visible en sitio')
        elif datos_prod is None:
            motivos.append('Sin datos de producto')
        if datos_prod and not datos_prod.get('ShowWithoutStock', False):
            motivos.append('ShowWithoutStock desactivado')
        sc_list = datos_sku.get('SalesChannels', [])
        sc_ids = set(sc_list) if sc_list else set()
        if not any(sc in sc_ids for sc in sales_channels_filtro):
            canales_str = ', '.join(str(sc) for sc in sales_channels_filtro)
            motivos.append(f'Sin ninguno de SC [{canales_str}]')
        # Precio y stock solo se verifican si fueron solicitados
        if incluir_precio_stock:
            if precio is None:
                motivos.append('Sin precio')
            if stock is not None and stock <= 0:
                motivos.append('Sin stock')
        activo = len(motivos) == 0
        return activo, ', '.join(motivos)

    # -- Precio y stock -------------------------------------------------------

    def _obtener_precio(self, sku_id: int, contexto: _ContextoVtex) -> float | None:
        url = f"{contexto.url_base_seller}/api/pricing/prices/{sku_id}"
        datos = self._request_con_retry(url, contexto.session_seller, silenciar_404=True)
        if datos:
            return datos.get('basePrice')
        return None

    def _obtener_stock(self, sku_id: int, contexto: _ContextoVtex) -> int | None:
        url = f"{contexto.url_base_seller}/api/logistics/pvt/inventory/skus/{sku_id}"
        datos = self._request_con_retry(url, contexto.session_seller, silenciar_404=True)
        if datos:
            balance = datos.get('balance', [])
            return sum(
                max(a.get('totalQuantity', 0) - a.get('reservedQuantity', 0), 0)
                for a in balance
            )
        return None

    # -- HTTP con reintentos --------------------------------------------------

    def _request_con_retry(
        self, url: str, session: requests.Session, silenciar_404: bool = False,
        info_error: list | None = None,
    ) -> dict | list | None:
        """Hace GET con reintentos. Si se pasa info_error (lista), appendea el motivo del fallo."""
        for intento in range(1, REINTENTOS_MAXIMOS + 1):
            try:
                resp = session.get(url=url, timeout=30)
                if resp.status_code == 429:
                    wait = ESPERA_ENTRE_REINTENTOS * (2 ** (intento - 1)) + random.uniform(0, 1)
                    logger.warning(f"429 en {url}, esperando {wait:.1f}s ({intento}/{REINTENTOS_MAXIMOS})")
                    time.sleep(wait)
                    continue
                if resp.status_code == 404 and silenciar_404:
                    return None
                resp.raise_for_status()
                return resp.json()
            except requests.exceptions.HTTPError as e:
                if intento < REINTENTOS_MAXIMOS:
                    wait = ESPERA_ENTRE_REINTENTOS * (2 ** (intento - 1)) + random.uniform(0, 1)
                    time.sleep(wait)
                    continue
                logger.error(f"HTTP error {url}: {e}")
                if info_error is not None:
                    info_error.append(f"HTTP {resp.status_code}")
                return None
            except Exception as e:
                if intento < REINTENTOS_MAXIMOS:
                    wait = ESPERA_ENTRE_REINTENTOS * (2 ** (intento - 1)) + random.uniform(0, 1)
                    logger.warning(f"Error {url}: {e}, reintentando en {wait:.1f}s ({intento}/{REINTENTOS_MAXIMOS})")
                    time.sleep(wait)
                    continue
                logger.error(f"Error {url}: {e}")
                if info_error is not None:
                    info_error.append(str(e))
                return None
        # Agoto reintentos (tipicamente por 429 consecutivos)
        if info_error is not None:
            info_error.append(f"429 x{REINTENTOS_MAXIMOS}")
        return None

    # -- Utilidades -----------------------------------------------------------

    def _resultado_error(self, sku_id: int, motivo: str) -> dict:
        resultado = {col: '' for col in COLUMNAS_EXPORT}
        resultado.update({
            'ACTIVO': 'ERROR', 'FOTO': 'ERROR', 'CATALOGADO': 'ERROR',
            '_IDSKU': str(sku_id), 'Motivo': motivo, 'Precio': None, 'Stock': None,
        })
        return resultado

    def _generar_excel(
        self, tarea: TareaCatalogacion, resultados: list[dict],
        filas_imagenes: list[dict] | None = None,
    ) -> None:
        ahora = datetime.now()
        nombre = f'EXPORT_VTEX_{ahora.day}_{ahora.month}.xlsx'
        directorio_relativo = os.path.join('output', str(ahora.year), str(ahora.month), str(ahora.day))
        directorio = os.path.join(settings.MEDIA_ROOT, directorio_relativo)
        os.makedirs(directorio, exist_ok=True)
        ruta_final = os.path.join(directorio, nombre)

        wb = Workbook(write_only=True)

        # Solapa 1: Catalogacion
        ws = wb.create_sheet(title='Catalogacion')
        ws.append(COLUMNAS_EXPORT)
        for fila in resultados:
            valores = []
            for col in COLUMNAS_EXPORT:
                v = fila.get(col, '')
                if isinstance(v, str):
                    v = _limpiar_para_excel(v)
                valores.append(v)
            ws.append(valores)

        # Solapa 2: Imagenes
        if filas_imagenes:
            ws_img = wb.create_sheet(title='Imagenes')
            ws_img.append(COLUMNAS_IMAGENES)
            for fila in filas_imagenes:
                valores = []
                for col in COLUMNAS_IMAGENES:
                    v = fila.get(col, '')
                    if isinstance(v, str):
                        v = _limpiar_para_excel(v)
                    valores.append(v)
                ws_img.append(valores)

        wb.save(ruta_final)

        # Solo setear en memoria, se guarda a DB en _finalizar_estado
        tarea.archivo_resultado = os.path.join(directorio_relativo, nombre)
        self._log(tarea, f"Excel generado: {nombre}")

    # -- Estado temporal (archivo JSON en vez de DB) --------------------------

    def _iniciar_estado_temporal(self, tarea: TareaCatalogacion) -> None:
        """Crea el archivo JSON temporal para esta tarea."""
        directorio = os.path.join(settings.MEDIA_ROOT, 'tmp')
        os.makedirs(directorio, exist_ok=True)
        self._ruta_temporal = os.path.join(directorio, f'tarea_{tarea.id}.json')
        self._estado = tarea.estado
        self._logs = tarea.logs or ''
        self._progreso_actual = 0
        self._progreso_total = 0

    def _escribir_estado_temporal(self) -> None:
        """Escribe el estado actual a archivo JSON. Escritura atomica via rename."""
        datos = {
            'estado': self._estado,
            'logs': self._logs,
            'progreso_actual': self._progreso_actual,
            'progreso_total': self._progreso_total,
        }
        ruta_tmp = self._ruta_temporal + '.tmp'
        with open(ruta_tmp, 'w', encoding='utf-8') as f:
            json.dump(datos, f)
        os.replace(ruta_tmp, self._ruta_temporal)

    def _finalizar_estado(self, tarea: TareaCatalogacion) -> None:
        """Vuelca todo a DB en un solo save y elimina el archivo temporal."""
        tarea.estado = self._estado
        tarea.logs = self._logs
        tarea.progreso_actual = self._progreso_actual
        tarea.progreso_total = self._progreso_total
        tarea.save()
        try:
            os.remove(self._ruta_temporal)
        except OSError:
            pass

    def _actualizar_estado(self, tarea: TareaCatalogacion, estado: str) -> None:
        with self._lock:
            self._estado = estado
            self._escribir_estado_temporal()

    def _log(self, tarea: TareaCatalogacion, mensaje: str) -> None:
        with self._lock:
            if self._logs:
                self._logs += f"\n{mensaje}"
            else:
                self._logs = mensaje
            self._escribir_estado_temporal()

    def _set_progreso(self, total: int, actual: int) -> None:
        """Setea progreso total/actual y escribe al archivo temporal."""
        with self._lock:
            self._progreso_total = total
            self._progreso_actual = actual
            self._progreso_pendiente = 0
            self._escribir_estado_temporal()

    def _incrementar_progreso(self, tarea: TareaCatalogacion) -> None:
        with self._lock:
            self._progreso_pendiente += 1
            if self._progreso_pendiente >= self.INTERVALO_PROGRESO:
                self._progreso_actual += self._progreso_pendiente
                self._progreso_pendiente = 0
                self._escribir_estado_temporal()

    def _flush_progreso(self, tarea: TareaCatalogacion) -> None:
        """Guarda el progreso pendiente que no alcanzo el intervalo."""
        with self._lock:
            if self._progreso_pendiente > 0:
                self._progreso_actual += self._progreso_pendiente
                self._progreso_pendiente = 0
                self._escribir_estado_temporal()


# -- Funciones auxiliares de formato ------------------------------------------

def _si_no(valor) -> str:
    if valor is None:
        return ''
    if isinstance(valor, bool):
        return 'SI' if valor else 'NO'
    if isinstance(valor, str):
        return 'SI' if valor.lower() in ('true', 'sí', 'si', 'yes') else 'NO'
    return 'SI' if valor else 'NO'


def _num(valor) -> str:
    if valor is None:
        return ''
    return str(valor)


_ILLEGAL_XML_CHARS = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]')
_HTML_TAG = re.compile(r'<[^>]+>')


def _limpiar_para_excel(valor: str) -> str:
    if not valor:
        return valor
    texto = unescape(valor)
    texto = _HTML_TAG.sub(' ', texto)
    texto = _ILLEGAL_XML_CHARS.sub('', texto)
    texto = ' '.join(texto.split())
    return texto.strip()
